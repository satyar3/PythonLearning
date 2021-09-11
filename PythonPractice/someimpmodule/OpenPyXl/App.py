import openpyxl as xl

# For bar Chart
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell1 = sheet["A1"]
print(cell1.value)
cell2 = sheet.cell(1, 1)
print(cell2.value)

print(sheet.max_row)

# Reading Cell value
for i in range(1, sheet.max_row + 1):
    print(sheet.cell(i, 1).value)

# Setting up a cell value
for i in range(2, sheet.max_row + 1):
    cell_original = sheet.cell(i, 3)
    cell_corrected = sheet.cell(i, 5)
    cell_corrected_value = float(cell_original.value) * 0.90
    cell_corrected.value = cell_corrected_value

wb.save("transaction_upd.xlsx")

# Bar Chart
for i in range(2, sheet.max_row + 1):
    cell_original = sheet.cell(i, 3)
    cell_corrected = sheet.cell(i, 5)
    cell_corrected_value = float(cell_original.value) * 0.90
    cell_corrected.value = cell_corrected_value

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "G2")
wb.save("transaction_upd.xlsx")
